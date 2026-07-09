"""
Fyers Research Data Collector — Multi-Year Daily Data + Full Classification
Run: python fyers_research_data.py

Fetches N years of daily candles, classifies every day across multiple
research dimensions, and saves a formatted EXCEL file with:
  - "Data" sheet   : every day, color-coded green/red, with filters
  - "Summary" sheet: ready-made pivot-style count tables (live formulas)

REQUIREMENTS
------------
pip install pandas openpyxl
"""

from fyers_apiv3 import fyersModel
from datetime import datetime, timedelta
import sys
import os
import time

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule

sys.path.append(os.path.join(os.path.dirname(__file__), ".."))
from creditials import client_id

# --------------------------------------------------
# Access Token
# --------------------------------------------------
with open("access_token.txt", "r") as f:
    access_token = f.read().strip()

fyers = fyersModel.FyersModel(
    client_id=client_id, is_async=False, token=access_token, log_path=""
)

# --------------------------------------------------
# Config
# --------------------------------------------------
SYMBOL        = "NSE:NIFTY50-INDEX"
YEARS_BACK    = 5
CHUNK_DAYS    = 365
SLEEP_BETWEEN_CALLS = 0.4
OUTPUT_XLSX   = "research_data.xlsx"

TREND_WINDOW  = 20
VOL_WINDOW    = 20
GAP_THRESHOLD = 0.3

BANDS = [
    ("Strong Positive",   2.0,   float("inf")),
    ("Positive",          1.0,   2.0),
    ("Slightly Positive", 0.3,   1.0),
    ("Consolidation",    -0.3,   0.3),
    ("Slightly Negative", -1.0, -0.3),
    ("Negative",          -2.0, -1.0),
    ("Strong Negative",  float("-inf"), -2.0),
]

def classify_band(pct):
    for label, low, high in BANDS:
        if low < pct <= high:
            return label
    return "Consolidation"

# --------------------------------------------------
# Fetch (chunked for the 366-day API limit)
# --------------------------------------------------
def fetch_chunk(symbol, start, end):
    data = {
        "symbol": symbol, "resolution": "D", "date_format": "1",
        "range_from": start, "range_to": end, "cont_flag": "1", "oi_flag": "0",
    }
    response = fyers.history(data=data)
    if response.get("s") != "ok":
        print(f"  [ERROR {response.get('code','?')}] {response.get('message','Unknown error')} ({start} -> {end})")
        return []
    return response.get("candles", [])

def fetch_multi_year(symbol=SYMBOL, years_back=YEARS_BACK):
    now = datetime.now()
    market_close_today = now.replace(hour=15, minute=30, second=0, microsecond=0)
    overall_end = now if now >= market_close_today else now - timedelta(days=1)
    overall_start = overall_end - timedelta(days=years_back * 365)

    all_candles = []
    chunk_end = overall_end
    print(f"\n  Symbol : {symbol}")
    print(f"  Range  : {overall_start.date()} -> {overall_end.date()}  (~{years_back} years)")
    print("  Fetching in chunks (366-day API limit)...\n")

    while chunk_end > overall_start:
        chunk_start = max(chunk_end - timedelta(days=CHUNK_DAYS), overall_start)
        s_str, e_str = chunk_start.strftime("%Y-%m-%d"), chunk_end.strftime("%Y-%m-%d")
        print(f"    Fetching {s_str} -> {e_str} ...", end=" ")
        candles = fetch_chunk(symbol, s_str, e_str)
        print(f"{len(candles)} candles")
        all_candles.extend(candles)
        chunk_end = chunk_start - timedelta(days=1)
        time.sleep(SLEEP_BETWEEN_CALLS)

    unique = {c[0]: c for c in all_candles}
    return sorted(unique.values(), key=lambda c: c[0])

# --------------------------------------------------
# Build classified dataframe
# --------------------------------------------------
def build_research_table(candles):
    df = pd.DataFrame(candles, columns=["ts", "open", "high", "low", "close", "volume"])
    df["date"] = pd.to_datetime(df["ts"], unit="s")
    df["day"] = df["date"].dt.strftime("%A")

    df["pct_change"] = ((df["close"] - df["open"]) / df["open"]) * 100
    df["movement"] = df["pct_change"].apply(lambda x: "Positive" if x > 0 else ("Negative" if x < 0 else "Flat"))
    df["magnitude_band"] = df["pct_change"].apply(classify_band)

    df["prev_close"] = df["close"].shift(1)
    df["gap_pct"] = ((df["open"] - df["prev_close"]) / df["prev_close"]) * 100
    df["gap_type"] = df["gap_pct"].apply(
        lambda x: "N/A" if pd.isna(x) else ("Gap Up" if x > GAP_THRESHOLD else ("Gap Down" if x < -GAP_THRESHOLD else "No Gap"))
    )

    df["day_range_pct"] = ((df["high"] - df["low"]) / df["open"]) * 100
    avg_range = df["day_range_pct"].rolling(VOL_WINDOW, min_periods=5).mean()
    df["volatility_band"] = [
        "N/A" if pd.isna(a) else ("High" if r > a * 1.3 else ("Low" if r < a * 0.7 else "Normal"))
        for r, a in zip(df["day_range_pct"], avg_range)
    ]

    avg_vol = df["volume"].rolling(VOL_WINDOW, min_periods=5).mean()
    df["volume_type"] = [
        "N/A" if pd.isna(a) else ("High" if v > a * 1.3 else ("Low" if v < a * 0.7 else "Normal"))
        for v, a in zip(df["volume"], avg_vol)
    ]

    df["ma"] = df["close"].rolling(TREND_WINDOW, min_periods=5).mean()
    df["ma_prev"] = df["ma"].shift(5)
    df["trend"] = [
        "N/A" if pd.isna(m) or pd.isna(mp) else
        ("Uptrend" if (m - mp) / mp * 100 > 0.5 else ("Downtrend" if (m - mp) / mp * 100 < -0.5 else "Sideways"))
        for m, mp in zip(df["ma"], df["ma_prev"])
    ]

    streaks, current_streak, current_dir = [], 0, None
    for m in df["movement"]:
        if m == current_dir and m != "Flat":
            current_streak += 1
        else:
            current_streak, current_dir = 1, m
        streaks.append(current_streak if m != "Flat" else 0)
    df["streak"] = streaks

    df["date_str"] = df["date"].dt.strftime("%d-%b-%Y")
    out = df[[
        "date_str", "day", "open", "high", "low", "close", "volume",
        "pct_change", "movement", "magnitude_band",
        "gap_pct", "gap_type", "day_range_pct", "volatility_band",
        "volume_type", "trend", "streak",
    ]].rename(columns={"date_str": "date"})

    return out.round({"open": 2, "high": 2, "low": 2, "close": 2, "pct_change": 2, "gap_pct": 2, "day_range_pct": 2})

# --------------------------------------------------
# Write formatted Excel workbook
# --------------------------------------------------
HEADER_FILL = PatternFill("solid", start_color="1F4E78", end_color="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN_FILL  = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
RED_FILL    = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")
GREEN_FONT  = Font(color="006100")
RED_FONT    = Font(color="9C0006")
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_data_sheet(wb, df):
    ws = wb.active
    ws.title = "Data"

    for col_idx, col_name in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(df.itertuples(index=False), start=2):
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=val).border = BORDER

    last_row = len(df) + 1
    last_col_letter = get_column_letter(len(df.columns))
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{last_col_letter}{last_row}"

    movement_col = df.columns.get_loc("movement") + 1
    movement_letter = get_column_letter(movement_col)
    full_range = f"A2:{last_col_letter}{last_row}"
    ws.conditional_formatting.add(
        full_range,
        CellIsRule(operator="equal", formula=[f'"Positive"'], fill=GREEN_FILL, font=GREEN_FONT)
    ) if False else None
    # Apply row-color based on movement column value using formula-based rule
    from openpyxl.formatting.rule import FormulaRule
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'${movement_letter}2="Positive"'], fill=GREEN_FILL)
    )
    ws.conditional_formatting.add(
        full_range,
        FormulaRule(formula=[f'${movement_letter}2="Negative"'], fill=RED_FILL)
    )

    widths = [12, 11, 9, 9, 9, 9, 12, 11, 11, 16, 9, 10, 13, 15, 11, 10, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    return last_row, {name: get_column_letter(i + 1) for i, name in enumerate(df.columns)}

def write_summary_sheet(wb, last_row, col_letters):
    ws = wb.create_sheet("Summary")
    ws.column_dimensions["A"].width = 22
    for c in "BCDEFGH":
        ws.column_dimensions[c].width = 16

    def title(cell, text):
        ws[cell] = text
        ws[cell].font = Font(bold=True, size=12, color="1F4E78")

    def header(cell, text):
        ws[cell] = text
        ws[cell].fill = HEADER_FILL
        ws[cell].font = HEADER_FONT
        ws[cell].alignment = Alignment(horizontal="center")

    trend_col = col_letters["trend"]
    band_col = col_letters["magnitude_band"]
    gap_col = col_letters["gap_type"]
    move_col = col_letters["movement"]
    vola_col = col_letters["volatility_band"]
    day_col = col_letters["day"]
    r = f"2:{last_row}"

    # --- Table 1: Trend x Magnitude Band ---
    title("A1", "Trend vs Magnitude Band (day count)")
    bands = [b[0] for b in BANDS]
    trends = ["Uptrend", "Downtrend", "Sideways"]
    header("A2", "Trend")
    for i, b in enumerate(bands, start=2):
        header(f"{get_column_letter(i)}2", b)
    for row_i, tr in enumerate(trends, start=3):
        ws.cell(row=row_i, column=1, value=tr)
        for col_i, b in enumerate(bands, start=2):
            col_l = get_column_letter(col_i)
            ws.cell(row=row_i, column=col_i,
                    value=f'=COUNTIFS(Data!${trend_col}${r.split(":")[0]}:${trend_col}${r.split(":")[1]},$A{row_i},'
                          f'Data!${band_col}${r.split(":")[0]}:${band_col}${r.split(":")[1]},{col_l}$2)')

    # --- Table 2: Gap Type x Movement ---
    base_row = 3 + len(trends) + 2
    title(f"A{base_row}", "Gap Type vs Movement (day count)")
    header(f"A{base_row+1}", "Gap Type")
    for i, m in enumerate(["Positive", "Negative", "Flat"], start=2):
        header(f"{get_column_letter(i)}{base_row+1}", m)
    for row_i, gt in enumerate(["Gap Up", "Gap Down", "No Gap"], start=base_row+2):
        ws.cell(row=row_i, column=1, value=gt)
        for col_i, m in enumerate(["Positive", "Negative", "Flat"], start=2):
            col_l = get_column_letter(col_i)
            ws.cell(row=row_i, column=col_i,
                    value=f'=COUNTIFS(Data!${gap_col}$2:${gap_col}${last_row},$A{row_i},'
                          f'Data!${move_col}$2:${move_col}${last_row},{col_l}${base_row+1})')

    # --- Table 3: Volatility Band x Movement ---
    base_row2 = base_row + 2 + 3 + 2
    title(f"A{base_row2}", "Volatility Band vs Movement (day count)")
    header(f"A{base_row2+1}", "Volatility")
    for i, m in enumerate(["Positive", "Negative", "Flat"], start=2):
        header(f"{get_column_letter(i)}{base_row2+1}", m)
    for row_i, vb in enumerate(["High", "Normal", "Low"], start=base_row2+2):
        ws.cell(row=row_i, column=1, value=vb)
        for col_i, m in enumerate(["Positive", "Negative", "Flat"], start=2):
            col_l = get_column_letter(col_i)
            ws.cell(row=row_i, column=col_i,
                    value=f'=COUNTIFS(Data!${vola_col}$2:${vola_col}${last_row},$A{row_i},'
                          f'Data!${move_col}$2:${move_col}${last_row},{col_l}${base_row2+1})')

    # --- Table 4: Day of Week x Movement ---
    base_row3 = base_row2 + 2 + 3 + 2
    title(f"A{base_row3}", "Day of Week vs Movement (day count)")
    header(f"A{base_row3+1}", "Day")
    for i, m in enumerate(["Positive", "Negative", "Flat"], start=2):
        header(f"{get_column_letter(i)}{base_row3+1}", m)
    weekdays = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]
    for row_i, d in enumerate(weekdays, start=base_row3+2):
        ws.cell(row=row_i, column=1, value=d)
        for col_i, m in enumerate(["Positive", "Negative", "Flat"], start=2):
            col_l = get_column_letter(col_i)
            ws.cell(row=row_i, column=col_i,
                    value=f'=COUNTIFS(Data!${day_col}$2:${day_col}${last_row},$A{row_i},'
                          f'Data!${move_col}$2:${move_col}${last_row},{col_l}${base_row3+1})')

def build_workbook(df, out_path):
    wb = Workbook()
    last_row, col_letters = write_data_sheet(wb, df)
    write_summary_sheet(wb, last_row, col_letters)
    wb.save(out_path)

# --------------------------------------------------
# Entry Point
# --------------------------------------------------
if __name__ == "__main__":
    candles = fetch_multi_year(SYMBOL, YEARS_BACK)
    if not candles:
        print("  [!] No data fetched.")
    else:
        df = build_research_table(candles)
        build_workbook(df, OUTPUT_XLSX)
        print(f"\n  Saved {len(df)} rows -> {OUTPUT_XLSX}")
        print("  Open it in Excel: 'Data' tab has every day (green=up, red=down, filters on).")
        print("  'Summary' tab has ready-made count tables (Trend vs Band, Gap vs Movement, etc.)")
        print("  These use live formulas -- if you edit the Data tab, Summary updates automatically.\n")